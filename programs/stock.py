from selenium import webdriver
from selenium.webdriver.common.keys import keys
import time
import openpyxl
from pathlib import path 

wait_imp = 10
excel_path = path(r"Downloads\stock_data.xlsx")
wb = openpyxl.load_workbook(excel_path)
ws = wb["CMP"]

print ("step 1--> Reading Excel sheet, please wait.....")
s_row =4
c_list =[]
avg_val = []
qnty_list = []

while ws.cell(row = s_row,column = 2).value != None:

  c_name =ws.cell(row = s_row,column = 2).value
  val_1 = ws.cell(row=s_row,column = 4).value
  qnty = ws.cell(row = s_row,column =5).value
  c_list.append(c_name)
  avg_val.append(val_1)
  qnty_list.append(qnty)
  s_row +=1
print("company name available in database")
[print('   ->',name) for name in c_list]
time.sleep(2)
print('\n')

CO = webdriver.ChromeOptions()
CO.add_experimental_option('Use automation extension',False)

