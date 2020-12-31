from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl as xl
from pathlib import Path

wait_imp = 10
excel_path = Path(r'stock_data.xlsx')
print (excel_path)
wb = xl.load_workbook(excel_path)
ws = wb["CMP"]

# Read company name from excelsheet
print ("Step 1 --> Reading Excel-sheet, Please wait....")
s_row     = 2
c_list    = []
url_list = []
avg_val   = []
qnty_list = []
while ws.cell(row = s_row, column= 1).value != None:
    c_name = ws.cell(row = s_row, column= 1).value
    url = ws.cell(row = s_row, column= 2).value
    val_1 = ws.cell(row = s_row, column= 4).value
    qnty = ws.cell(row = s_row, column= 5).value
    c_list.append(c_name)
    avg_val.append(val_1)
    url_list.append(url)
    qnty_list.append(qnty)
    s_row += 1    
print ("Company name available in Database")
[print('    ->',name) for name in c_list]
time.sleep(2)
print ('\n')
# create a webdriver object for chrome-option and configure
CO = webdriver.ChromeOptions()
CO.add_experimental_option('useAutomationExtension', False)
CO.add_argument('--ignore-certificate-errors')
CO.add_argument('--start-maximized')
wd = webdriver.Chrome(r'C:\Users\user\Desktop\my code\chromedriver.exe',options=CO)

print ("Step 2 --> Opening Finance website\n")
wd.implicitly_wait(wait_imp)
#wd.get("https://www.moneycontrol.com")






time.sleep(5)
print ("******************************************************************************")
print ("                      Getting Live Stock Value !! Please wait ...\n")

print (c_list)
print (url_list)
for i in range(len(c_list)):
   # src = wd.find_element_by_id ("search_str")
   # src.send_keys(c_list[i])
    #src.send_keys(Keys.RETURN)
    #wd.implicitly_wait(wait_imp)
    wd.get(url_list[i])
    time.sleep(5)

    s_v = wd.find_element_by_xpath('//*[@id="nsespotval"]')
    print(s_v.get_attribute('value'))
    ws.cell(row=4+i, column= 3, value = s_v.get_attribute('value'))
    
    diff = (avg_val[i] - float(s_v.get_attribute('value').replace(',' , '')))* qnty_list[i]
    per_diff = (diff/(avg_val[i]*qnty_list[i]))*100
    print ("{:>23} -> CMP {:<7} Current P/L->[{:>8.2f}] %P/L -> {:>6.2f}%".format(c_list[i],s_v.get_attribute('value'), diff, per_diff))

print ('\n')
print ("Step 3 --> Writing Latest Price into Excel-sheet ....\n")
time.sleep(1)
wb.save(excel_path)

print ("Step 4 --> Successfully Written  \n")
print ("Step 5 --> Closing browser !\n")
print (" ----------------------- FINISHED !! ------------------------")

time.sleep(1)
wd.close()